import csv
import io
from datetime import datetime
from models import db, User, Attendance, BusRoute, BusPass

def parse_attendance_csv_or_excel(file_stream, filename, marked_by_id):
    """
    Parses a CSV or Excel (XLSX) file containing student attendance.
    Expected Columns: student_email, date, status (Present/Absent), route_number_or_id
    Returns: (success_count, error_messages)
    """
    records = []
    errors = []
    
    # Check if Excel
    if filename.endswith('.xlsx'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_stream)
            sheet = wb.active
            # Read header
            header = [cell.value for cell in sheet[1]]
            required = {'student_email', 'date', 'status'}
            if not required.issubset(set(header)):
                return 0, [f"Missing required columns. Found headers: {header}"]
            
            # Map headers to indices
            email_idx = header.index('student_email')
            date_idx = header.index('date')
            status_idx = header.index('status')
            route_idx = header.index('route_number_or_id') if 'route_number_or_id' in header else -1
            
            for row_num in range(2, sheet.max_row + 1):
                row = [sheet.cell(row=row_num, column=col_num).value for col_num in range(1, len(header) + 1)]
                if not any(row):
                    continue
                records.append({
                    'email': row[email_idx],
                    'date': row[date_idx],
                    'status': row[status_idx],
                    'route_identifier': row[route_idx] if route_idx != -1 else None
                })
        except ImportError:
            return 0, ["openpyxl library is required to read .xlsx files. Please upload as .csv instead or install openpyxl."]
        except Exception as e:
            return 0, [f"Error reading Excel sheet: {e}"]
    else:
        # CSV parsing
        try:
            # Wrap standard text wrapper around binary file stream if needed
            if isinstance(file_stream, bytes):
                text_stream = io.StringIO(file_stream.decode('utf-8'))
            else:
                text_stream = io.StringIO(file_stream.read().decode('utf-8'))
            
            reader = csv.DictReader(text_stream)
            if not reader.fieldnames:
                return 0, ["CSV file is empty or missing header row."]
            
            required = {'student_email', 'date', 'status'}
            if not required.issubset(set(reader.fieldnames)):
                return 0, [f"Missing required columns. Found: {reader.fieldnames}"]
            
            for row in reader:
                records.append({
                    'email': row.get('student_email'),
                    'date': row.get('date'),
                    'status': row.get('status'),
                    'route_identifier': row.get('route_number_or_id')
                })
        except Exception as e:
            return 0, [f"Error reading CSV file: {e}"]
            
    # Process the parsed records
    success_count = 0
    from werkzeug.security import generate_password_hash
    
    for idx, rec in enumerate(records, start=2):
        email = rec['email']
        date_str = rec['date']
        status = rec['status']
        route_ident = rec['route_identifier']
        
        if not email or not date_str or not status:
            errors.append(f"Row {idx}: Missing email, date, or status value.")
            continue
            
        # Standardize email
        email = email.strip().lower()
        status = status.strip().capitalize()
        
        # Verify student exists
        student = User.query.filter_by(email=email, role='student').first()
        if not student:
            errors.append(f"Row {idx}: Student email '{email}' not found in database.")
            continue
            
        # Parse date
        parsed_date = None
        if isinstance(date_str, datetime):
            parsed_date = date_str.date()
        elif hasattr(date_str, 'date'):  # date object
            parsed_date = date_str
        else:
            # Attempt to parse string dates
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%d/%m/%Y'):
                try:
                    parsed_date = datetime.strptime(str(date_str).strip(), fmt).date()
                    break
                except ValueError:
                    continue
            if not parsed_date:
                errors.append(f"Row {idx}: Invalid date format '{date_str}'. Use YYYY-MM-DD.")
                continue
                
        # Resolve route
        route = None
        if route_ident:
            route = BusRoute.query.filter(
                (BusRoute.id == route_ident) | (BusRoute.bus_number == str(route_ident))
            ).first()
        
        # Fallback to student's active pass route, or just first route
        if not route:
            # Let's see if student has a pass and we can map to a default route
            active_pass = BusPass.query.filter_by(student_id=student.id).first()
            route = BusRoute.query.first()
            if not route:
                errors.append(f"Row {idx}: No routes available in system to log attendance.")
                continue
                
        # Standardize status
        if status not in ('Present', 'Absent'):
            status = 'Present' if 'pres' in status.lower() or 'p' == status.lower() else 'Absent'
            
        # Check if record already exists for student on this date
        existing = Attendance.query.filter_by(student_id=student.id, date=parsed_date).first()
        if existing:
            existing.status = status
            existing.route_id = route.id
            existing.marked_by = marked_by_id
        else:
            new_att = Attendance(
                student_id=student.id,
                route_id=route.id,
                date=parsed_date,
                status=status,
                marked_by=marked_by_id
            )
            db.session.add(new_att)
            
        success_count += 1
        
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        errors.append(f"Database error while saving attendance: {e}")
        return 0, errors
        
    return success_count, errors


def parse_student_enrollment_csv_or_excel(file_stream, filename):
    """
    Parses a CSV or Excel (XLSX) file containing student enrollments.
    Expected Columns: name, email, phone (optional)
    Returns: (success_count, error_messages)
    """
    records = []
    errors = []
    
    if filename.endswith('.xlsx'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_stream)
            sheet = wb.active
            header = [cell.value for cell in sheet[1]]
            required = {'name', 'email'}
            if not required.issubset(set(header)):
                return 0, [f"Missing required columns. Found headers: {header}"]
            
            name_idx = header.index('name')
            email_idx = header.index('email')
            phone_idx = header.index('phone') if 'phone' in header else -1
            
            for row_num in range(2, sheet.max_row + 1):
                row = [sheet.cell(row=row_num, column=col_num).value for col_num in range(1, len(header) + 1)]
                if not any(row):
                    continue
                records.append({
                    'name': row[name_idx],
                    'email': row[email_idx],
                    'phone': row[phone_idx] if phone_idx != -1 else None
                })
        except ImportError:
            return 0, ["openpyxl library is required to read .xlsx files. Please upload as .csv instead."]
        except Exception as e:
            return 0, [f"Error reading Excel sheet: {e}"]
    else:
        try:
            if isinstance(file_stream, bytes):
                text_stream = io.StringIO(file_stream.decode('utf-8'))
            else:
                text_stream = io.StringIO(file_stream.read().decode('utf-8'))
            
            reader = csv.DictReader(text_stream)
            if not reader.fieldnames:
                return 0, ["CSV file is empty or missing header row."]
            
            required = {'name', 'email'}
            if not required.issubset(set(reader.fieldnames)):
                return 0, [f"Missing required columns. Found: {reader.fieldnames}"]
            
            for row in reader:
                records.append({
                    'name': row.get('name'),
                    'email': row.get('email'),
                    'phone': row.get('phone')
                })
        except Exception as e:
            return 0, [f"Error reading CSV file: {e}"]
            
    success_count = 0
    from werkzeug.security import generate_password_hash
    
    for idx, rec in enumerate(records, start=2):
        name = rec['name']
        email = rec['email']
        phone = rec['phone']
        
        if not name or not email:
            errors.append(f"Row {idx}: Missing name or email.")
            continue
            
        email = email.strip().lower()
        name = name.strip()
        phone = str(phone).strip() if phone else None
        
        # Verify student doesn't exist
        existing = User.query.filter_by(email=email).first()
        if existing:
            errors.append(f"Row {idx}: User with email '{email}' already exists. Skipping.")
            continue
            
        # Create student with default hashed password 'csmss123'
        hashed = generate_password_hash('csmss123')
        new_student = User(
            name=name,
            email=email,
            password_hash=hashed,
            role='student',
            phone=phone
        )
        db.session.add(new_student)
        success_count += 1
        
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        errors.append(f"Database error while saving students: {e}")
        return 0, errors
        
    return success_count, errors
